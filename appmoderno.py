import os
import sqlite3
import shutil
from datetime import datetime
from functools import partial
import hashlib
import re

import qrcode
from kivy.uix.image import Image
import openpyxl
from fpdf import FPDF
from kivy.animation import Animation
from kivy.app import App
from kivy.core.text import LabelBase
from kivy.core.window import Window
from kivy.graphics import Color, Ellipse, Rectangle, RoundedRectangle
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.uix.screenmanager import ScreenManager, Screen, FadeTransition
from kivy.utils import get_color_from_hex
from kivy.clock import Clock
from pyzbar import pyzbar
from PIL import Image as PilImage
from kivy.uix.camera import Camera
from kivy_garden.graph import Graph, BarPlot

# --- Paleta de Cores (FOCO NO VERDE) ---
colors = {
    "background": get_color_from_hex("#E8F5E9"), "primary": get_color_from_hex("#2E7D32"),
    "secondary": get_color_from_hex("#37474F"), "accent": get_color_from_hex("#4CAF50"),
    "white": get_color_from_hex("#FFFFFF"), "success": get_color_from_hex("#388E3C"),
    "danger": get_color_from_hex("#D32F2F"), "warning": get_color_from_hex("#FFA000"),
}
LabelBase.register(name="MaterialSymbols", fn_regular="MaterialSymbolsOutlined-Regular.ttf")
DB_FILE = "floricultura.db"

def conectar(): return sqlite3.connect(DB_FILE)

def inicializar_banco():
    if not os.path.exists('images'): os.makedirs('images')
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS estoque (codigo TEXT PRIMARY KEY, nome TEXT NOT NULL, tipo TEXT, quantidade INTEGER NOT NULL, imagem_path TEXT)')
    try: cursor.execute("ALTER TABLE estoque ADD COLUMN imagem_path TEXT")
    except sqlite3.OperationalError: pass
    cursor.execute('CREATE TABLE IF NOT EXISTS log (id INTEGER PRIMARY KEY AUTOINCREMENT, data TEXT, acao TEXT)')
    cursor.execute('CREATE TABLE IF NOT EXISTS usuarios (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL)')
    cursor.execute("SELECT COUNT(id) FROM usuarios")
    if cursor.fetchone()[0] == 0:
        password_hash = hashlib.sha256('admin'.encode('utf-8')).hexdigest()
        cursor.execute("INSERT INTO usuarios (username, password_hash) VALUES (?, ?)", ('admin', password_hash))
    conn.commit()
    conn.close()

def adicionar_log(acao):
    conn = conectar(); cursor = conn.cursor(); data = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    cursor.execute("INSERT INTO log (data, acao) VALUES (?, ?)", (data, acao)); conn.commit(); conn.close()

class ModernButton(Button):
    def __init__(self, **kwargs):
        radius_val = kwargs.pop('radius', 12); bg_color = kwargs.pop('bg_color', colors['accent'])
        if isinstance(radius_val, int): radius = [radius_val]
        else: radius = radius_val
        super().__init__(**kwargs)
        self.background_color = (0, 0, 0, 0); self.background_normal = ''; self.background_down = ''; self.markup = True
        with self.canvas.before:
            Color(0, 0, 0, 0.2); self.shadow = RoundedRectangle(radius=radius)
            self.bg_color_instruction = Color(rgba=bg_color); self.rect = RoundedRectangle(radius=radius)
        self.bind(pos=self.update_rect, size=self.update_rect); self.bind(state=self.on_state)
    def update_rect(self, *args):
        self.rect.pos = self.pos; self.rect.size = self.size
        self.shadow.pos = (self.pos[0] + 2, self.pos[1] - 2); self.shadow.size = self.size
    def on_state(self, instance, value):
        if value == 'down': self.shadow.pos = self.pos
        else: self.shadow.pos = (self.pos[0] + 2, self.pos[1] - 2)

class FabWithBadge(FloatLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.size_hint = (None, None); self.size = (76, 76)
        self.fab = ModernButton(text="[font=MaterialSymbols]shopping_cart[/font]", size_hint=(None, None), size=(56, 56), pos_hint={'center_x': 0.5, 'center_y': 0.5}, bg_color=colors['warning'], radius=28)
        self.add_widget(self.fab)
        with self.canvas:
            self.badge_color = Color(rgba=colors['danger']); self.badge_circle = Ellipse(size=(22, 22), pos=(self.width - 24, self.height - 24))
        self.badge_label = Label(text="", font_size='12sp', size_hint=(None,None), size=(22,22), pos=(self.width-24, self.height-24)); self.add_widget(self.badge_label)
        self.update_badge(0)
    def update_badge(self, count):
        self.badge_label.text = str(count)
        if count > 0: self.badge_color.a = 1; self.badge_label.color[3] = 1
        else: self.badge_color.a = 0; self.badge_label.color[3] = 0

class MenuLateral(BoxLayout):
    def __init__(self, **kwargs):
        self.app = App.get_running_app()
        super().__init__(**kwargs)
        self.orientation = 'vertical'; self.size_hint = (None, None); self.width = 240
        self.height = Window.height; self.x = -self.width; self.y = 0; self.spacing = 15; self.padding = [15, 15]; self.aberto = False
        with self.canvas.before:
            Color(rgba=colors['primary']); self.rect = Rectangle(size=self.size, pos=self.pos)
        self.bind(pos=self.update_rect, size=self.update_rect)
        header = BoxLayout(size_hint_y=None, height=50, spacing=10)
        btn_voltar = Button(text="[font=MaterialSymbols]arrow_back[/font]", markup=True, on_press=lambda x: self.toggle(), size_hint=(None, None), size=(50,50), background_color=(0,0,0,0))
        header.add_widget(btn_voltar); header.add_widget(Label(text="Menu", font_size='18sp', size_hint_x=0.8)); self.add_widget(header)
        menu_items = [
            ("dashboard", "Dashboard", "dashboard"),
            ("list", "Estoque", "estoque"),
            ("photo_library", "Catálogo", "catalogo"),
            ("qr_code_scanner", "Ler QR Code", "scanner"),
            ("add_circle", "Adicionar Item", self.app.abrir_popup_adicionar),
            ("save", "Exportar Excel", self.app.exportar_excel),
            ("picture_as_pdf", "Exportar PDF", self.app.exportar_pdf),
            ("history", "Ver Log", self.app.ver_log),
        ]
        for icon, text, action in menu_items:
            if isinstance(action, str): on_press_action = partial(self.mudar_tela, action)
            else: on_press_action = partial(self.executar_acao, action)
            btn = ModernButton(text=f"[font=MaterialSymbols]{icon}[/font]  {text}", on_press=on_press_action, size_hint_y=None, height=50, bg_color=colors['accent'])
            self.add_widget(btn)
        self.add_widget(Label())
        btn_sair = ModernButton(text="[font=MaterialSymbols]logout[/font]  Sair", on_press=self.app.logout, size_hint_y=None, height=50, bg_color=colors['danger'])
        self.add_widget(btn_sair)
    def update_rect(self, *args): self.rect.pos = self.pos; self.rect.size = self.size
    def toggle(self): anim = Animation(x=0 if not self.aberto else -self.width, d=0.3, t='out_quad').start(self); self.aberto = not self.aberto
    def mudar_tela(self, nome_tela, instance):
        if self.aberto: self.toggle()
        self.app.mudar_tela_principal(nome_tela)
    def executar_acao(self, acao, instance):
        if self.aberto: self.toggle()
        acao(instance)

class ListItemCard(BoxLayout):
    def __init__(self, **kwargs):
        self.item_data = kwargs.pop('item_data', {})
        super().__init__(**kwargs)
        self.main_app = App.get_running_app(); self.orientation = 'horizontal'; self.size_hint_y = None; self.height = 70; self.padding = [10,]
        with self.canvas.before:
            Color(rgba=colors['white']); self.rect = RoundedRectangle(radius=[12,], size=self.size, pos=self.pos)
        self.bind(pos=self.update_rect, size=self.update_rect)
        info_layout = BoxLayout(orientation='vertical', padding=[5,0])
        self.name_label = Label(text=f"[b]{self.item_data.get('nome', '')}[/b] ({self.item_data.get('codigo', '')})", color=colors['secondary'], markup=True, halign='left', valign='middle')
        self.stock_label = Label(text=f"Em estoque: {self.item_data.get('quantidade', 0)}", color=colors['secondary'], halign='left', valign='middle')
        def update_text_size(instance, width): self.name_label.text_size = (width, None); self.stock_label.text_size = (width, None)
        info_layout.bind(width=update_text_size)
        info_layout.add_widget(self.name_label); info_layout.add_widget(self.stock_label)
        actions_layout = BoxLayout(size_hint_x=None, width=230, spacing=5)
        qr_button = Button(text="[font=MaterialSymbols]qr_code_2[/font]", on_press=self.generate_qr, markup=True, background_color=(0,0,0,0), size_hint=(None, None), size=(50,50), color=colors['secondary'])
        add_cart_button = Button(text="[font=MaterialSymbols]add_shopping_cart[/font]", on_press=self.add_to_cart, markup=True, background_color=(0,0,0,0), size_hint=(None, None), size=(50,50), color=colors['success'])
        edit_button = Button(text="[font=MaterialSymbols]edit[/font]", on_press=self.edit_item, markup=True, background_color=(0,0,0,0), size_hint=(None, None), size=(50,50), color=colors['accent'])
        delete_button = Button(text="[font=MaterialSymbols]delete[/font]", on_press=self.delete_item, markup=True, background_color=(0,0,0,0), size_hint=(None, None), size=(50,50), color=colors['danger'])
        actions_layout.add_widget(qr_button); actions_layout.add_widget(add_cart_button); actions_layout.add_widget(edit_button); actions_layout.add_widget(delete_button)
        self.add_widget(info_layout); self.add_widget(actions_layout)
    def generate_qr(self, instance): self.main_app.gerar_e_mostrar_qr_code(self.item_data)
    def add_to_cart(self, instance): self.main_app.adicionar_item_ao_carrinho_pela_lista(self.item_data)
    def edit_item(self, instance): self.main_app.abrir_popup_editar(self.item_data)
    def delete_item(self, instance): self.main_app.abrir_popup_confirmar_remocao(self.item_data)
    def update_rect(self, *args): self.rect.pos = self.pos; self.rect.size = self.size

class CartItemWidget(BoxLayout):
    def __init__(self, **kwargs):
        item_data = kwargs.pop('item_data', {}); plus_callback = kwargs.pop('plus_callback', None); minus_callback = kwargs.pop('minus_callback', None)
        super().__init__(**kwargs)
        self.orientation = 'horizontal'; self.size_hint_y = None; self.height = 50; self.padding = [10, 0]
        with self.canvas.before:
            Color(rgba=colors['accent']); self.bg_rect = RoundedRectangle(radius=[12,]); self.bind(pos=self.update_rect, size=self.update_rect)
        self.add_widget(Label(text=f"{item_data.get('nome', '')} ({item_data.get('codigo', '')})", color=colors['white'], halign='left', valign='middle', text_size=(self.width*0.8, None)))
        qty_layout = BoxLayout(size_hint_x=None, width=120)
        btn_minus = ModernButton(text="-", on_press=minus_callback, size_hint_x=None, width=40, bg_color=colors['primary'])
        lbl_qty = Label(text=str(item_data.get('qtd', 0)), size_hint_x=None, width=40, color=colors['white'], bold=True)
        btn_plus = ModernButton(text="+", on_press=plus_callback, size_hint_x=None, width=40, bg_color=colors['primary'])
        qty_layout.add_widget(btn_minus); qty_layout.add_widget(lbl_qty); qty_layout.add_widget(btn_plus)
        self.add_widget(qty_layout)
    def update_rect(self, *args): self.bg_rect.pos = self.pos; self.bg_rect.size = self.size

class CatalogoCard(BoxLayout):
    def __init__(self, **kwargs):
        item_data = kwargs.pop('item_data', {})
        super().__init__(**kwargs)
        self.orientation = 'vertical'; self.size_hint_y = None; self.height = 240; self.padding = 8
        with self.canvas.before:
            Color(rgba=colors['white']); self.bg_rect = RoundedRectangle(radius=[15,], size=self.size, pos=self.pos)
        self.bind(pos=self.update_rect, size=self.update_rect)
        caminho_imagem = item_data.get('imagem_path') or 'placeholder.png'
        if not os.path.exists(caminho_imagem):
            caminho_imagem = 'placeholder.png'
            if not os.path.exists('placeholder.png'):
                img = PilImage.new('RGB', (200, 200), color = (220, 220, 220)); d = PilImage.Draw(img)
                d.text((10,10), "Sem Imagem", fill=(100,100,100)); img.save('placeholder.png')
        self.add_widget(Image(source=caminho_imagem, nocache=True, size_hint_y=0.7))
        self.add_widget(Label(text=f"[b]{item_data.get('nome', '')}[/b]", markup=True, color=colors['secondary'], size_hint_y=0.15))
        self.add_widget(Label(text=f"Estoque: {item_data.get('quantidade', 0)}", color=colors['secondary'], font_size='12sp', size_hint_y=0.15))
    def update_rect(self, *args): self.bg_rect.pos = self.pos; self.bg_rect.size = self.size

class LoginScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        with self.canvas.before:
            Color(rgba=colors['background']); self.bg = Rectangle(pos=self.pos, size=Window.size)
        layout_principal = FloatLayout()
        self.login_box = BoxLayout(orientation='vertical', padding=40, spacing=20, size_hint=(None, None), size=(380, 420), pos_hint={'center_x': 0.5, 'center_y': 0.5})
        with self.login_box.canvas.before:
            Color(0, 0, 0, 0.2); self.shadow = RoundedRectangle(radius=[15,])
            Color(rgba=colors['white']); self.rect_bg = RoundedRectangle(radius=[15,])
        self.login_box.bind(pos=self.update_rect_bg, size=self.update_rect_bg)
        title = Label(text="[b]Bem-vindo![/b]", markup=True, font_size='32sp', color=colors['primary'], size_hint_y=None, height=50)
        subtitle = Label(text="Acesse seu controle de estoque", font_size='16sp', color=colors['secondary'], size_hint_y=None, height=20)
        self.username = TextInput(hint_text="Usuário", multiline=False, size_hint_y=None, height=55, font_size='16sp', padding=[15, (55-18)/2], background_color=colors['background'], foreground_color=colors['secondary'], cursor_color=colors['primary'], background_normal='', background_active='')
        self.password = TextInput(hint_text="Senha", password=True, multiline=False, size_hint_y=None, height=55, font_size='16sp', padding=[15, (55-18)/2], background_color=colors['background'], foreground_color=colors['secondary'], cursor_color=colors['primary'], background_normal='', background_active='')
        self.feedback = Label(text="", color=colors['danger'], size_hint_y=None, height=20)
        login_button = ModernButton(text="Entrar", on_press=self.verificar_login, size_hint_y=None, height=55, bg_color=colors['primary'], radius=12)
        self.login_box.add_widget(title); self.login_box.add_widget(subtitle); self.login_box.add_widget(BoxLayout(size_hint_y=None, height=20))
        self.login_box.add_widget(self.username); self.login_box.add_widget(self.password)
        self.login_box.add_widget(self.feedback); self.login_box.add_widget(login_button)
        layout_principal.add_widget(self.login_box); self.add_widget(layout_principal)
        self.login_box.pos_hint = {'center_x': 0.5, 'center_y': 1.5}
        anim = Animation(pos_hint={'center_x': 0.5, 'center_y': 0.5}, d=0.8, t='out_bounce'); anim.start(self.login_box)
    def update_rect_bg(self, instance, value):
        self.rect_bg.pos = instance.pos; self.rect_bg.size = instance.size
        self.shadow.pos = (instance.pos[0] - 4, instance.pos[1] - 4); self.shadow.size = instance.size
    def verificar_login(self, instance):
        user = self.username.text; pwd = self.password.text
        if not user or not pwd: self.feedback.text = "Por favor, preencha todos os campos."; return
        conn = conectar(); cursor = conn.cursor(); pwd_hash = hashlib.sha256(pwd.encode('utf-8')).hexdigest()
        cursor.execute("SELECT id FROM usuarios WHERE username = ? AND password_hash = ?", (user, pwd_hash))
        if cursor.fetchone():
            self.feedback.text = ""; App.get_running_app().on_login_success()
        else: self.feedback.text = "Usuário ou senha incorreta."
        conn.close()

class EstoqueScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout_estoque = BoxLayout(orientation='vertical', padding=[0, 10, 10, 10], spacing=10)
        self.barra_busca = TextInput(hint_text="Buscar plantas...", multiline=False, size_hint_y=None, height=50, padding=[15, (50-18)/2, 0, 0], background_color=colors['white'], foreground_color=colors['secondary'], font_size='16sp')
        self.barra_busca.bind(text=self.atualizar_lista)
        self.lista_estoque = GridLayout(cols=1, spacing=10, size_hint_y=None)
        self.lista_estoque.bind(minimum_height=self.lista_estoque.setter('height'))
        scroll_estoque = ScrollView(bar_width=10, bar_color=colors['primary']); scroll_estoque.add_widget(self.lista_estoque)
        layout_estoque.add_widget(self.barra_busca); layout_estoque.add_widget(scroll_estoque)
        self.add_widget(layout_estoque)
    def on_enter(self, *args): self.atualizar_lista()
    def atualizar_lista(self, *args):
        filtro = self.barra_busca.text.lower(); self.lista_estoque.clear_widgets(); conn = conectar(); cursor = conn.cursor()
        cursor.row_factory = sqlite3.Row; query = "SELECT * FROM estoque"
        params = []
        if filtro:
            query += " WHERE lower(nome) LIKE ? OR lower(tipo) LIKE ? OR lower(codigo) LIKE ?"; params = [f"%{filtro}%", f"%{filtro}%", f"%{filtro}%"]
        query += " ORDER BY nome"; cursor.execute(query, params)
        for row in cursor.fetchall(): self.lista_estoque.add_widget(ListItemCard(item_data=dict(row)))
        conn.close()

class CatalogoScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout_catalogo = BoxLayout(orientation='vertical', padding=[0, 10, 10, 10], spacing=10)
        self.grid_catalogo = GridLayout(cols=3, spacing=10, size_hint_y=None); self.grid_catalogo.bind(minimum_height=self.grid_catalogo.setter('height'))
        scroll_catalogo = ScrollView(bar_width=10, bar_color=colors['primary']); scroll_catalogo.add_widget(self.grid_catalogo)
        layout_catalogo.add_widget(scroll_catalogo)
        self.add_widget(layout_catalogo)
    def on_enter(self, *args): self.atualizar_catalogo()
    def atualizar_catalogo(self, *args):
        self.grid_catalogo.clear_widgets(); conn = conectar(); cursor = conn.cursor()
        cursor.row_factory = sqlite3.Row; cursor.execute("SELECT * FROM estoque ORDER BY nome")
        for row in cursor.fetchall(): self.grid_catalogo.add_widget(CatalogoCard(item_data=dict(row)))
        conn.close()

class DashboardScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=[0, 10, 10, 10], spacing=20)
        estoque_box = BoxLayout(orientation='vertical', size_hint_y=0.4)
        estoque_box.add_widget(Label(text="Alertas de Estoque Baixo (<= 5 Itens)", bold=True, color=colors['secondary'], size_hint_y=None, height=30))
        scroll_estoque = ScrollView()
        self.low_stock_grid = GridLayout(cols=1, spacing=5, size_hint_y=None)
        self.low_stock_grid.bind(minimum_height=self.low_stock_grid.setter('height'))
        scroll_estoque.add_widget(self.low_stock_grid); estoque_box.add_widget(scroll_estoque); layout.add_widget(estoque_box)
        vendas_box = BoxLayout(orientation='vertical', size_hint_y=0.6)
        vendas_box.add_widget(Label(text="Top 5 Plantas Mais Vendidas", bold=True, color=colors['secondary'], size_hint_y=None, height=30))
        self.graph = Graph(xlabel='Plantas', ylabel='Qtd. Vendida', x_ticks_minor=0, x_ticks_major=1, y_ticks_major=5, y_grid_label=True, x_grid_label=True, padding=5, x_grid=False, y_grid=True, xmin=-1, xmax=5, ymin=0)
        vendas_box.add_widget(self.graph); layout.add_widget(vendas_box); self.add_widget(layout)
    def on_enter(self, *args): self.atualizar_alerta_estoque(); self.atualizar_grafico_vendas()
    def atualizar_alerta_estoque(self):
        self.low_stock_grid.clear_widgets(); conn = conectar(); cursor = conn.cursor()
        cursor.execute("SELECT nome, quantidade FROM estoque WHERE quantidade <= 5 ORDER BY quantidade ASC")
        itens_baixos = cursor.fetchall(); conn.close()
        if not itens_baixos: self.low_stock_grid.add_widget(Label(text="Nenhum item com estoque baixo.", color=colors['success'])); return
        for nome, qtd in itens_baixos: self.low_stock_grid.add_widget(Label(text=f"[b]{nome}[/b] - Apenas {qtd} unidades!", markup=True, color=colors['danger']))
    def atualizar_grafico_vendas(self):
        conn = conectar(); cursor = conn.cursor(); cursor.execute("SELECT acao FROM log WHERE acao LIKE 'VENDA REALIZADA:%'"); vendas_log = cursor.fetchall(); conn.close()
        contagem_vendas = {}
        for (log,) in vendas_log:
            itens_vendidos = re.findall(r" (\d+)x '([^']+)';", log)
            for qtd, nome in itens_vendidos:
                nome = nome.strip()
                if nome in contagem_vendas: contagem_vendas[nome] += int(qtd)
                else: contagem_vendas[nome] = int(qtd)
        top_5_vendidos = sorted(contagem_vendas.items(), key=lambda item: item[1], reverse=True)[:5]
        for plot in self.graph.plots: self.graph.remove_plot(plot)
        if not top_5_vendidos: self.graph.ylabel = "Nenhuma venda registrada"; self.graph.x_ticks_label = []; self.graph.xmax = 5; return
        self.graph.ylabel = "Qtd. Vendida"
        nomes = [item[0] for item in top_5_vendidos]; quantidades = [item[1] for item in top_5_vendidos]
        plot_points = [(i, qtd) for i, qtd in enumerate(quantidades)]
        self.graph.x_ticks_label = [Label(text='\n'.join(nome.split()), shorten=True, max_lines=2) for nome in nomes]
        self.graph.ymax = max(quantidades) + 1 if quantidades else 10; self.graph.xmax = len(plot_points)
        plot = BarPlot(color=colors['accent'], bar_spacing=0.2); plot.points = plot_points; self.graph.add_plot(plot)

class ScannerScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation='vertical')
        try:
            self.camera = Camera(play=False, resolution=(640, 480))
            self.layout.add_widget(self.camera)
        except Exception as e:
            self.layout.add_widget(Label(text=f"Erro ao iniciar a câmera:\n{e}\n\nVerifique permissões e se o provedor (OpenCV) está instalado.", halign='center'))
        self.add_widget(self.layout)
    def on_enter(self, *args):
        if hasattr(self, 'camera'): self.camera.play = True; self.schedule = Clock.schedule_interval(self.decode_qr_code, 1.0)
    def on_leave(self, *args):
        if hasattr(self, 'camera'): self.camera.play = False
        if hasattr(self, 'schedule'): self.schedule.cancel()
    def decode_qr_code(self, dt):
        try:
            texture = self.camera.texture
            if not texture: return
            pil_image = PilImage.frombytes(mode='RGBA', size=texture.size, data=texture.pixels)
            decoded_objects = pyzbar.decode(pil_image)
            if decoded_objects:
                self.on_leave()
                App.get_running_app().show_plant_from_qr(decoded_objects[0].data.decode('utf-8'))
        except Exception as e: print(f"Erro ao decodificar QR Code: {e}")

class PlantDetailScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation='vertical', padding=20, spacing=15)
        header = BoxLayout(size_hint_y=None, height=50)
        back_button = ModernButton(text="[font=MaterialSymbols]arrow_back[/font]", on_press=self.go_back, size_hint_x=None, width=55, bg_color=colors['primary'])
        self.plant_name_label = Label(text="Detalhes da Planta", font_size='24sp', color=colors['primary'], bold=True)
        header.add_widget(back_button); header.add_widget(self.plant_name_label)
        self.plant_image = Image(size_hint_y=0.6, nocache=True)
        self.plant_code_label = Label(font_size='18sp', color=colors['secondary'], size_hint_y=0.1, markup=True)
        self.plant_stock_label = Label(font_size='18sp', color=colors['secondary'], size_hint_y=0.1, markup=True)
        self.plant_type_label = Label(font_size='18sp', color=colors['secondary'], size_hint_y=0.1, markup=True)
        self.layout.add_widget(header); self.layout.add_widget(self.plant_image); self.layout.add_widget(self.plant_code_label)
        self.layout.add_widget(self.plant_stock_label); self.layout.add_widget(self.plant_type_label)
        self.add_widget(self.layout)
    def update_details(self, plant_data):
        self.plant_name_label.text = plant_data.get('nome', '')
        self.plant_code_label.text = f"Código: [b]{plant_data.get('codigo', '')}[/b]"
        self.plant_stock_label.text = f"Em Estoque: [b]{plant_data.get('quantidade', 0)}[/b] unidades"
        self.plant_type_label.text = f"Tipo: [b]{plant_data.get('tipo', '')}[/b]"
        caminho_imagem = plant_data.get('imagem_path') or 'placeholder.png'
        if not os.path.exists(caminho_imagem): caminho_imagem = 'placeholder.png'
        self.plant_image.source = caminho_imagem; self.plant_image.reload()
    def go_back(self, instance):
        app = App.get_running_app()
        app.mudar_tela_principal('estoque')

class AppLayout(FloatLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.content_sm = ScreenManager(transition=FadeTransition(duration=0.15))
        self.add_widget(self.content_sm)
        
        self.content_sm.add_widget(DashboardScreen(name='dashboard'))
        self.content_sm.add_widget(EstoqueScreen(name='estoque'))
        self.content_sm.add_widget(CatalogoScreen(name='catalogo'))
        
        header = BoxLayout(orientation='horizontal', size_hint=(1, None), height=50, pos_hint={'top': 1}, padding=[10,0,10,0])
        self.menu = MenuLateral()
        menu_button = ModernButton(text="[font=MaterialSymbols]menu[/font]", size_hint=(None, 1), width=50, on_press=lambda x: self.menu.toggle(), bg_color=colors['primary'])
        self.header_title = Label(text="", font_size='20sp', color=colors['secondary'], bold=True)
        header.add_widget(menu_button); header.add_widget(self.header_title)
        
        self.add_widget(header)
        self.add_widget(self.menu)
        self.fab = FabWithBadge(pos_hint={'right': 0.98, 'y': 0.02})
        self.fab.fab.bind(on_press=App.get_running_app().abrir_popup_venda)
        self.add_widget(self.fab)

        self.content_sm.bind(on_current_screen=self.update_header_title)
        self.content_sm.current = 'dashboard'

    def update_header_title(self, instance, screen):
        titles = {
            'dashboard': 'Dashboard', 'estoque': 'Controle de Estoque',
            'catalogo': 'Catálogo de Plantas', 'scanner': 'Ler QR Code',
            'plant_detail': 'Detalhes da Planta'
        }
        self.header_title.text = titles.get(screen.name, '')

class EstoqueApp(App):
    def build(self):
        Window.clearcolor = colors['background']
        inicializar_banco()
        self.carrinho = {}
        
        self.root = ScreenManager(transition=FadeTransition(duration=0.2))
        self.root.add_widget(LoginScreen(name='login'))
        
        main_app_screen = Screen(name='main_app')
        main_app_screen.add_widget(AppLayout())
        self.root.add_widget(main_app_screen)

        # Telas que não fazem parte do layout principal ficam no gerenciador raiz
        self.root.add_widget(ScannerScreen(name='scanner'))
        self.root.add_widget(PlantDetailScreen(name='plant_detail'))
        
        self.root.current = 'login'
        return self.root

    def on_login_success(self): 
        self.mudar_tela_principal('dashboard')
        self.root.current = 'main_app'
        
    def logout(self, instance):
        self.carrinho.clear()
        main_layout = self.root.get_screen('main_app').children[0]
        main_layout.fab.update_badge(0)
        self.root.get_screen('login').feedback.text = ""
        self.root.current = 'login'

    def gerar_e_mostrar_qr_code(self, item_data):
        # Cria o conteúdo para o QR Code (pode ser um link para um futuro site)
        qr_content = f"https://pyflora.app?id={item_data.get('codigo', '')}"
        
        # Gera a imagem do QR Code
        qr_img = qrcode.make(qr_content)
        temp_path = "temp_qr_code.png"
        qr_img.save(temp_path)

        # Cria o layout do popup
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        content.add_widget(Label(text=f"QR Code para: [b]{item_data.get('nome', 'N/A')}[/b]", markup=True, size_hint_y=None, height=40, color=colors['secondary']))
        content.add_widget(Image(source=temp_path, nocache=True)) # nocache=True é importante para atualizar a imagem
        
        popup = self._criar_popup(
            title="QR Code Gerado", 
            content=content, 
            size_hint=(None, None), # Desativa o size_hint para usar tamanho fixo
            size=(400, 450) # Tamanho fixo para o popup do QR
        )
        
        # Remove o arquivo temporário quando o popup for fechado
        popup.bind(on_dismiss=lambda x: os.remove(temp_path))
        popup.open()
        
    def mudar_tela_principal(self, nome_tela):
        """Função única para navegar para qualquer tela."""
        # Se for uma tela de conteúdo, navega o SM interno
        if nome_tela in ['estoque', 'catalogo', 'dashboard']:
            self.root.current = 'main_app'
            main_layout = self.root.get_screen('main_app').children[0]
            main_layout.content_sm.current = nome_tela
        # Se for uma tela global, navega o SM raiz
        else:
            self.root.current = nome_tela

    def show_plant_from_qr(self, qr_data):
        try: plant_code = qr_data.split('?id=')[1].strip()
        except IndexError: return
        conn = conectar(); cursor = conn.cursor(); cursor.row_factory = sqlite3.Row
        cursor.execute("SELECT * FROM estoque WHERE codigo = ?", (plant_code,)); row = cursor.fetchone(); conn.close()
        if row:
            plant_data = dict(row)
            detail_screen = self.root.get_screen('plant_detail')
            detail_screen.update_details(plant_data)
            self.root.current = 'plant_detail'
        else:
            self.mudar_tela_principal('estoque')
            Popup(title="Erro", content=Label(text=f"Planta com código '{plant_code}'\nnão foi encontrada."), size_hint=(0.7, 0.4)).open()

    def adicionar_item_ao_carrinho_pela_lista(self, item_data):
        codigo = item_data['codigo']; conn = conectar(); cursor = conn.cursor(); cursor.execute("SELECT quantidade FROM estoque WHERE codigo = ?", (codigo,)); resultado = cursor.fetchone(); conn.close()
        if not resultado: return
        qtd_estoque = resultado[0]; qtd_no_carrinho = self.carrinho.get(codigo, {}).get('qtd', 0)
        if qtd_estoque <= qtd_no_carrinho:
            Popup(title="Estoque Insuficiente", content=Label(text=f"Não há mais unidades de\n'{item_data['nome']}'."), size_hint=(0.6, 0.3)).open(); return
        if codigo in self.carrinho: self.carrinho[codigo]['qtd'] += 1
        else: self.carrinho[codigo] = {'nome': item_data['nome'], 'qtd': 1, 'estoque_max': qtd_estoque}
        self.root.get_screen('main_app').children[0].fab.update_badge(len(self.carrinho))

    def _criar_popup(self, title, content, **kwargs):
        # Unifica a criação de popups, aceitando tanto size_hint quanto size
        config = {'title': title, 'content': content, 'separator_color': colors['primary'], 'background_color': colors['white']}
        if 'size_hint' not in kwargs and 'size' not in kwargs:
            config['size_hint'] = (0.9, 0.9)
        config.update(kwargs)
        return Popup(**config)
    
    def abrir_popup_venda(self, instance):
        if not self.carrinho:
            Popup(title="Carrinho Vazio", content=Label(text="Adicione itens pela lista principal primeiro."), size_hint=(0.5, 0.3)).open(); return
        content = BoxLayout(orientation='vertical', padding=10, spacing=10); carrinho_layout = GridLayout(cols=1, spacing=8, size_hint_y=None)
        carrinho_layout.bind(minimum_height=carrinho_layout.setter('height')); carrinho_scroll = ScrollView(); carrinho_scroll.add_widget(carrinho_layout)
        popup = self._criar_popup(title="Revisar e Finalizar Venda", content=content)
        def aumentar_qtd_carrinho(codigo, instance):
            if codigo in self.carrinho and self.carrinho[codigo]['qtd'] < self.carrinho[codigo]['estoque_max']: self.carrinho[codigo]['qtd'] += 1; atualizar_carrinho_ui()
        def diminuir_qtd_carrinho(codigo, instance):
            if codigo in self.carrinho:
                self.carrinho[codigo]['qtd'] -= 1
                if self.carrinho[codigo]['qtd'] <= 0:
                    del self.carrinho[codigo]; self.root.get_screen('main_app').children[0].fab.update_badge(len(self.carrinho))
                    if not self.carrinho: popup.dismiss()
                atualizar_carrinho_ui()
        def atualizar_carrinho_ui():
            carrinho_layout.clear_widgets()
            if not self.carrinho: popup.dismiss(); return
            for codigo, item in self.carrinho.items():
                cart_item = CartItemWidget(item_data={'codigo': codigo, **item},plus_callback=partial(aumentar_qtd_carrinho, codigo),minus_callback=partial(diminuir_qtd_carrinho, codigo))
                carrinho_layout.add_widget(cart_item)
        def finalizar_venda(instance):
            if not self.carrinho: popup.dismiss(); return
            conn = conectar(); cursor = conn.cursor(); log_venda = "VENDA REALIZADA:"
            try:
                for codigo, item in self.carrinho.items():
                    cursor.execute("UPDATE estoque SET quantidade = quantidade - ? WHERE codigo = ?", (item['qtd'], codigo)); log_venda += f" {item['qtd']}x '{item['nome']}';"
                conn.commit(); adicionar_log(log_venda)
            except Exception as e: conn.rollback()
            finally: conn.close()
            self.carrinho.clear()
            main_layout = self.root.get_screen('main_app').children[0]
            main_layout.fab.update_badge(0)
            main_layout.content_sm.current_screen.on_enter()
            popup.dismiss()
        content.add_widget(carrinho_scroll)
        finalizar_btn = ModernButton(text="[font=MaterialSymbols]check_circle[/font] Finalizar Venda", on_press=finalizar_venda, size_hint_y=None, height=50, bg_color=colors['success'])
        content.add_widget(finalizar_btn); atualizar_carrinho_ui(); popup.open()
    
    def _abrir_popup_item(self, item_data=None):
        title = "Adicionar Novo Item" if item_data is None else "Editar Item"; content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        nome = TextInput(hint_text="Nome da Flor", text=item_data.get('nome', '') if item_data else "")
        tipo = TextInput(hint_text="Tipo (ex: Flor)", text=item_data.get('tipo', '') if item_data else "")
        quantidade = TextInput(hint_text="Quantidade", input_filter="int", text=str(item_data.get('quantidade', '')) if item_data else "")
        imagem_selecionada_path = item_data.get('imagem_path') if item_data else None
        imagem_label = Label(text=f"Imagem: {os.path.basename(imagem_selecionada_path or 'Nenhuma')}", size_hint_y=None, height=30, color=colors['secondary'])
        def abrir_seletor_imagem(instance):
            file_chooser_layout = BoxLayout(orientation='vertical'); user_path = os.path.expanduser('~'); file_chooser = FileChooserIconView(path=user_path)
            select_button = Button(text="Selecionar", size_hint_y=None, height=44)
            file_chooser_layout.add_widget(file_chooser); file_chooser_layout.add_widget(select_button)
            chooser_popup = Popup(title="Escolha uma imagem", content=file_chooser_layout, size_hint=(0.9, 0.9))
            def selecionar_arquivo(instance):
                nonlocal imagem_selecionada_path
                if file_chooser.selection: imagem_selecionada_path = file_chooser.selection[0]; imagem_label.text = f"Imagem: {os.path.basename(imagem_selecionada_path)}"
                chooser_popup.dismiss()
            select_button.bind(on_press=selecionar_arquivo); chooser_popup.open()
        btn_selecionar_img = ModernButton(text="Selecionar Imagem", on_press=abrir_seletor_imagem, size_hint_y=None, height=44)
        content.add_widget(nome); content.add_widget(tipo); content.add_widget(quantidade); content.add_widget(btn_selecionar_img); content.add_widget(imagem_label)
        popup = self._criar_popup(title=title, content=content, size_hint=(0.8, 0.7))
        def salvar(_):
            if not nome.text or not quantidade.text: return
            novo_path_imagem = None
            if imagem_selecionada_path and os.path.exists(imagem_selecionada_path) and (not item_data or imagem_selecionada_path != item_data.get('imagem_path')):
                nome_arquivo = f"{nome.text.replace(' ', '_').lower()}_{datetime.now().strftime('%f')}{os.path.splitext(imagem_selecionada_path)[1]}"
                novo_path_imagem = os.path.join('images', nome_arquivo); shutil.copy(imagem_selecionada_path, novo_path_imagem)
            conn = conectar(); cursor = conn.cursor()
            if item_data is None:
                cursor.execute("SELECT codigo FROM estoque ORDER BY codigo DESC LIMIT 1"); ultimo_codigo = cursor.fetchone()
                novo_id = int(ultimo_codigo[0][2:]) + 1 if ultimo_codigo else 1; codigo = f"FL{novo_id:03d}"
                cursor.execute("INSERT INTO estoque (codigo, nome, tipo, quantidade, imagem_path) VALUES (?, ?, ?, ?, ?)",(codigo, nome.text, tipo.text or 'Flor', int(quantidade.text), novo_path_imagem))
                adicionar_log(f"ADICIONOU: {quantidade.text}x '{nome.text}' ({codigo})")
            else:
                codigo = item_data['codigo']; path_final = novo_path_imagem or item_data.get('imagem_path')
                cursor.execute("UPDATE estoque SET nome=?, tipo=?, quantidade=?, imagem_path=? WHERE codigo=?",(nome.text, tipo.text, int(quantidade.text), path_final, codigo))
                adicionar_log(f"EDITOU: Item {codigo} para '{nome.text}'")
            conn.commit(); conn.close()
            main_layout = self.root.get_screen('main_app').children[0]
            main_layout.content_sm.current_screen.on_enter()
            popup.dismiss()
        content.add_widget(ModernButton(text="[font=MaterialSymbols]save[/font] Salvar", on_press=salvar, size_hint_y=None, height=44, bg_color=colors['success']))
        popup.open()
    def abrir_popup_adicionar(self, instance=None): self._abrir_popup_item(None)
    def abrir_popup_editar(self, item_data): self._abrir_popup_item(item_data)
    def abrir_popup_confirmar_remocao(self, item_data):
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        msg = f"Tem certeza que deseja remover o item\n[b]{item_data['nome']} ({item_data['codigo']})[/b]?"; content.add_widget(Label(text=msg, markup=True, color=colors['secondary']))
        popup = self._criar_popup("Confirmar Remoção", content, size_hint=(0.7, 0.4))
        def confirmar_remocao(_):
            conn = conectar(); cursor = conn.cursor(); cursor.execute("SELECT imagem_path FROM estoque WHERE codigo = ?", (item_data['codigo'],)); path = cursor.fetchone()
            if path and path[0] and os.path.exists(path[0]): os.remove(path[0])
            cursor.execute("DELETE FROM estoque WHERE codigo = ?", (item_data['codigo'],)); conn.commit(); conn.close()
            adicionar_log(f"REMOVEU: Item {item_data['nome']} ({item_data['codigo']})")
            main_layout = self.root.get_screen('main_app').children[0]
            main_layout.content_sm.current_screen.on_enter()
            popup.dismiss()
        btn_layout = BoxLayout(spacing=10, size_hint_y=None, height=44)
        btn_layout.add_widget(ModernButton(text="Cancelar", on_press=popup.dismiss, bg_color=colors['secondary'])); btn_layout.add_widget(ModernButton(text="Sim, Remover", on_press=confirmar_remocao, bg_color=colors['danger']))
        content.add_widget(btn_layout); popup.open()
    def exportar_excel(self, instance=None):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Estoque"; ws.append(["Código", "Nome", "Tipo", "Quantidade"]); conn = conectar(); cursor = conn.cursor()
        cursor.execute("SELECT codigo, nome, tipo, quantidade FROM estoque ORDER BY nome")
        for linha in cursor.fetchall(): ws.append(linha)
        conn.close(); filepath = "estoque_exportado.xlsx"; wb.save(filepath); adicionar_log("EXPORTOU para Excel")
        Popup(title="Sucesso", content=Label(text=f"Exportado para\n{os.path.abspath(filepath)}"), size_hint=(0.7, 0.3)).open()
    def exportar_pdf(self, instance=None):
        pdf = FPDF(); pdf.add_page(); pdf.set_font("Arial", 'B', size=16); pdf.cell(0, 10, "Relatório de Estoque", ln=True, align='C'); pdf.ln(10)
        pdf.set_font("Arial", 'B', size=12); pdf.cell(30, 10, 'Código', 1); pdf.cell(90, 10, 'Nome', 1); pdf.cell(40, 10, 'Tipo', 1); pdf.cell(30, 10, 'Qtd.', 1); pdf.ln()
        pdf.set_font("Arial", size=10); conn = conectar(); cursor = conn.cursor()
        cursor.execute("SELECT codigo, nome, tipo, quantidade FROM estoque ORDER BY nome")
        for codigo, nome, tipo, quantidade in cursor.fetchall():
            pdf.cell(30, 10, str(codigo), 1); pdf.cell(90, 10, nome.encode('latin-1', 'replace').decode('latin-1'), 1); pdf.cell(40, 10, tipo.encode('latin-1', 'replace').decode('latin-1'), 1); pdf.cell(30, 10, str(quantidade), 1); pdf.ln()
        conn.close(); filepath = "estoque_exportado.pdf"; pdf.output(filepath); adicionar_log("EXPORTOU para PDF")
        Popup(title="Sucesso", content=Label(text=f"Exportado para\n{os.path.abspath(filepath)}"), size_hint=(0.7, 0.3)).open()
    def ver_log(self, instance=None):
        content = BoxLayout(orientation='vertical', padding=10)
        with content.canvas.before:
            Color(rgba=colors['secondary']); self.rect_log = Rectangle(size=content.size, pos=content.pos)
        def update_rect_log(instance, value): self.rect_log.pos = instance.pos; self.rect_log.size = instance.size
        content.bind(pos=update_rect_log, size=update_rect_log)
        log_grid = GridLayout(cols=1, size_hint_y=None, spacing=5); log_grid.bind(minimum_height=log_grid.setter('height'))
        conn = conectar(); cursor = conn.cursor(); cursor.execute("SELECT data, acao FROM log ORDER BY id DESC LIMIT 100")
        for data, acao in cursor.fetchall():
            log_grid.add_widget(Label(text=f"[{data}] {acao}", size_hint_y=None, height=30, color=colors['white']))
        conn.close(); scroll = ScrollView(); scroll.add_widget(log_grid); content.add_widget(scroll)
        popup = Popup(title="Histórico de Movimentações", content=content, size_hint=(0.9, 0.8), background_color=colors['secondary'], separator_color=colors['primary'])
        popup.open()

if __name__ == '__main__':
    EstoqueApp().run()